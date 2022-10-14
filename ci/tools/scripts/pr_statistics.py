import argparse
import codecs
import csv
import datetime
import logging
import openpyxl
import os
import pandas as pd
import requests
import smtplib
import subprocess
import sys
import yaml
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from logging import handlers
from openpyxl.styles import Alignment, Border, PatternFill, Side
from xlsx2html import xlsx2html


class Logger(object):
    level_relations = {
        'debug': logging.DEBUG,
        'info': logging.INFO,
        'warning': logging.WARNING,
        'error': logging.ERROR,
        'crit': logging.CRITICAL
    }

    def __init__(self, filename, level='info', when='D', backCount=3,
                 fmt='%(asctime)s - %(pathname)s[line:%(lineno)d] - %(levelname)s: %(message)s'):
        self.logger = logging.getLogger(filename)
        format_str = logging.Formatter(fmt)  # set logging formatter
        self.logger.setLevel(self.level_relations.get(level))  # set logging level
        sh = logging.StreamHandler()  # output to the console
        sh.setFormatter(format_str)  # format the display on the screen
        th = handlers.TimedRotatingFileHandler(filename=filename, when=when, backupCount=backCount, encoding='utf-8')
        th.setFormatter(format_str)  # set the format written in the file
        self.logger.addHandler(sh)
        self.logger.addHandler(th)


log = Logger('statistics.log', level='debug')


def get_open_pr_list(full_name, token, page):
    """
    Get all open Pull Request of a repo
    :param full_name: name contains organization and repo name
    :param token: access_token of gitee account
    :param page: page
    :return: a list of open Pull Request information
    """
    open_pr_list = []
    url = 'https://gitee.com/api/v5/repos/{}/pulls/'.format(full_name)
    params = {
        'access_token': token,
        'state': 'open',
        'sort': 'created',
        'direction': 'asc',
        'page': page,
        'per_page': 100
    }
    r = requests.get(url, params=params)
    if r.status_code != 200:
        log.logger.error('Fail to get open pull requests list of repo {}'.format(full_name))
        return open_pr_list
    else:
        open_pr_list += r.json()
        if len(r.json()) == 100:
            get_open_pr_list(full_name, token, page + 1)
    log.logger.info('Get open_pr_list of repo {}'.format(full_name))
    return open_pr_list


def prepare_env():
    """
    Prepare repository and directory
    """
    log.logger.info('=' * 25 + ' PREPARE ENVIRONMENT ' + '=' * 25)
    if os.path.exists('community'):
        subprocess.call('rm -rf community', shell=True)
    subprocess.call('git clone https://gitee.com/openeuler/community.git', shell=True)
    if not os.path.exists('community'):
        log.logger.error('Fail to clone code, exit...')
        sys.exit(1)
    data_dir = 'data'
    if os.path.exists(data_dir):
        subprocess.call('rm -rf {}'.format(data_dir), shell=True)
    subprocess.call('mkdir data', shell=True)
    if not os.path.exists('data'):
        log.logger.error('Fail to make data directory, exit...')
        sys.exit(1)
    log.logger.info('ENV is already.\n')
    return data_dir


def get_sigs():
    """
    Get relationship between sigs and repositories
    """
    log.logger.info('=' * 25 + ' GET SIGS INFO ' + '=' * 25)
    sig_path = os.path.join('community', 'sig')
    sigs = []
    for i in sorted(os.listdir(sig_path)):
        if i in ['README.md', 'sig-template', 'sig-recycle', 'create_sig_info_template.py']:
            continue
        if i not in [x['name'] for x in sigs]:
            sigs.append({'name': i, 'repositories': []})
        if 'openeuler' in os.listdir(os.path.join(sig_path, i)):
            for filesdir, _, repos in os.walk(os.path.join(sig_path, i, 'openeuler')):
                for repo in repos:
                    for sig in sigs:
                        if sig['name'] == i:
                            repositories = sig['repositories']
                            repositories.append(os.path.join('openeuler', repo.split('.yaml')[0]))
        if 'src-openeuler' in os.listdir(os.path.join(sig_path, i)):
            for filesdir, _, src_repos in os.walk(os.path.join(sig_path, i, 'src-openeuler')):
                for src_repo in src_repos:
                    for sig in sigs:
                        if sig['name'] == i:
                            repositories = sig['repositories']
                            repositories.append(os.path.join('src-openeuler', src_repo.split('.yaml')[0]))
    log.logger.info('Get sigs info.\n')
    return sigs


def get_maintainers(sig):
    """
    Get maintainers of the sig and mark where "maintainers" come from
    :param sig: sig name
    :return: maintainers, sig_info_mark
    """
    owners_file = os.path.join('community', 'sig', sig, 'OWNERS')
    sig_info_file = os.path.join('community', 'sig', sig, 'sig-info.yaml')
    if os.path.exists(owners_file):
        with open(owners_file, 'r', encoding='utf-8') as f:
            maintainers = yaml.load(f.read(), Loader=yaml.Loader)['maintainers']
            return maintainers, False
    elif os.path.exists(sig_info_file):
        with open(sig_info_file, 'r', encoding='utf-8') as f:
            sig_info = yaml.load(f.read(), Loader=yaml.Loader)
            maintainers = [x['gitee_id'] for x in sig_info['maintainers']]
            return maintainers, True
    else:
        log.logger.error('ERROR! Find SIG {} has neither OWNERS file nor sig-info.yaml.'.format(sig))
        sys.exit(1)


def get_committers_mapping(sig):
    """
    Get mappings between repos and committers
    :param sig: sig name
    :return: committers_mapping
    """
    sig_info_file = os.path.join('community', 'sig', sig, 'sig-info.yaml')
    if not os.path.exists(sig_info_file):
        return []
    with open(sig_info_file, 'r', encoding='utf-8') as f:
        sig_info = yaml.load(f.read(), Loader=yaml.Loader)
    repositories = sig_info['repositories']
    committers_mapping = {}
    for i in repositories:
        if 'committers' in i.keys():
            repos = i['repo']
            committers = [x['gitee_id'] for x in i['committers']]
            for repo in repos:
                committers_mapping[repo] = committers
    return committers_mapping


def get_repo_members(maintainers, committers_mapping, repo):
    """
    Get reviewers of a repo
    :param maintainers: maintainers of the sig
    :param committers_mapping: mappings between repos and committers
    :param repo: full name of repo
    :return: reviewers
    """
    if repo not in committers_mapping.keys():
        return maintainers
    reviewers = committers_mapping[repo]
    return reviewers


def count_duration(start_time):
    """
    Count open days of a Pull Request by its start_time
    :param start_time: time when the Pull Request starts
    :return: duration in days
    """
    start_time = start_time.split('+')[0]
    today = datetime.datetime.today()
    start_date = datetime.datetime.strptime(start_time, '%Y-%m-%dT%H:%M:%S')
    duration = str((today - start_date).days)
    return duration


def create_email_mappings():
    """
    Generate mappings between gitee_id and email addresses
    """
    email_mappings = {}
    if not os.path.exists('community'):
        subprocess.call('git clone https://gitee.com/openeuler/community.git', shell=True)
    sig_path = os.path.join('community', 'sig')
    for i in sorted(os.listdir(sig_path)):
        if i in ['README.md', 'sig-template', 'sig-recycle', 'create_sig_info_template.py']:
            continue
        log.logger.info('Starting to get email mappings of sig {}'.format(i))
        owners_file = os.path.join(sig_path, i, 'OWNERS')
        sig_info_file = os.path.join(sig_path, i, 'sig-info.yaml')
        if os.path.exists(owners_file):
            f = open(owners_file, 'r', encoding='utf-8')
            maintainers = yaml.safe_load(f)['maintainers']
            f.close()
            for maintainer in maintainers:
                if maintainer not in email_mappings.keys():
                    email_mappings[maintainer] = ''
        if os.path.exists(sig_info_file):
            f = open(sig_info_file, 'r', encoding='utf-8')
            sig_info = yaml.safe_load(f)
            f.close()
            maintainers = sig_info['maintainers']
            for maintainer in maintainers:
                maintainer_gitee_id = maintainer['gitee_id']
                maintainer_email = maintainer.get('email')
                if maintainer_email in ['null', 'NA'] or not maintainer_email:
                    maintainer_email = ''
                email_mappings[maintainer_gitee_id] = maintainer_email
            repositories = sig_info.get('repositories')
            if not repositories:
                continue
            for r in repositories:
                if 'committers' in r.keys():
                    commtters = r['committers']
                    for committer in commtters:
                        committer_gitee_id = committer['gitee_id']
                        committer_email = committer['email']
                        if committer_email in ['null', 'NA'] or not committer_email:
                            committer_email = ''
                        email_mappings[committer_gitee_id] = committer_email
    # generate email_mappings.yaml
    with open('email_mapping.yaml', 'w', encoding='utf-8') as f:
        yaml.dump(email_mappings, f, default_flow_style=False)


def get_email_mappings():
    """
    Get email_mappings
    :return: email_mappings
    """
    create_email_mappings()
    if not os.path.exists('email_mapping.yaml'):
        log.logger.error('ERROR! Fail to generate email_mappings.')
        return {}
    email_mappings = yaml.safe_load(open('email_mapping.yaml'))
    return email_mappings


def csv_to_xlsx(filepath):
    """
    Convert a csv file to a xlsx file
    :param filepath: path of the csv file
    :return: path of the xlsx file
    """
    if not filepath.endswith('.csv'):
        return
    # sorting
    df = pd.read_csv(filepath, encoding='utf-8')
    data = df.sort_values(by='PR开启天数', ascending=False)
    data.to_csv(filepath, mode='w', index=False)

    csv_file = pd.read_csv(filepath, encoding='utf-8')
    xlsx_filepath = filepath.replace('.csv', '.xlsx')
    csv_file.to_excel(xlsx_filepath, sheet_name='open_pull_requests_statistics')
    if not os.path.exists(xlsx_filepath):
        log.logger.error('ERROR! Fail to generate {}'.format(xlsx_filepath))
        sys.exit(1)
    log.logger.info('Generate {}'.format(filepath.replace('.csv', '.xlsx')))
    return xlsx_filepath


def generate_one_row_table(xlsx_file):
    """
    Generate html file with only one line of form
    :param xlsx_file: path of the xlsx file
    """
    csv_file = xlsx_file.replace('.xlsx', '.csv')
    html_file = xlsx_file.replace('.xlsx', '.html')
    f1 = open(csv_file, 'r', encoding='utf-8')
    info = f1.readlines()[1]
    f1.close()
    sig, repo, ref, link, status, duration, reviewers = info.replace('\n', '').split(',', 6)
    template = """
    <!DOCTYPE html> <html lang="en"> <head> <meta charset="UTF-8"> <title>Title</title> </head> <body> 
    <table style="border-collapse: collapse" border="0" cellspacing="0" cellpadding="0"><colgroup></colgroup> <tr> 
    <td id="open_pull_requests_statistics!A1" style="border-bottom-color: #000000;border-bottom-style: 
    solid;border-bottom-width: 1px;border-collapse: collapse;border-left-color: #000000;border-left-style: 
    solid;border-left-width: 1px;border-right-color: #000000;border-right-style: solid;border-right-width: 
    1px;border-top-color: #000000;border-top-style: solid;border-top-width: 1px;font-size: Nonepx;font-weight: 
    bold;height: 19pt;text-align: center">SIG组</td> <td id="open_pull_requests_statistics!B1" 
    style="border-bottom-color: #000000;border-bottom-style: solid;border-bottom-width: 1px;border-collapse: 
    collapse;border-left-color: #000000;border-left-style: solid;border-left-width: 1px;border-right-color: 
    #000000;border-right-style: solid;border-right-width: 1px;border-top-color: #000000;border-top-style: 
    solid;border-top-width: 1px;font-size: Nonepx;font-weight: bold;height: 19pt;text-align: center">仓库</td> <td 
    id="open_pull_requests_statistics!C1" style="border-bottom-color: #000000;border-bottom-style: 
    solid;border-bottom-width: 1px;border-collapse: collapse;border-left-color: #000000;border-left-style: 
    solid;border-left-width: 1px;border-right-color: #000000;border-right-style: solid;border-right-width: 
    1px;border-top-color: #000000;border-top-style: solid;border-top-width: 1px;font-size: Nonepx;font-weight: 
    bold;height: 19pt;text-align: center">目标分支</td> <td 
    id="open_pull_requests_statistics!D1" style="border-bottom-color: #000000;border-bottom-style: 
    solid;border-bottom-width: 1px;border-collapse: collapse;border-left-color: #000000;border-left-style: 
    solid;border-left-width: 1px;border-right-color: #000000;border-right-style: solid;border-right-width: 
    1px;border-top-color: #000000;border-top-style: solid;border-top-width: 1px;font-size: Nonepx;font-weight: 
    bold;height: 19pt;text-align: center">PR链接</td><td id="open_pull_requests_statistics!E1" 
    style="border-bottom-color: #000000;border-bottom-style: solid;border-bottom-width: 1px;border-collapse: 
    collapse;border-left-color: #000000;border-left-style: solid;border-left-width: 1px;border-right-color: 
    #000000;border-right-style: solid;border-right-width: 1px;border-top-color: #000000;border-top-style: 
    solid;border-top-width: 1px;font-size: Nonepx;font-weight: bold;height: 19pt;text-align: center">PR状态</td> <td 
    id="open_pull_requests_statistics!F1" style="border-bottom-color: #000000;border-bottom-style: 
    solid;border-bottom-width: 1px;border-collapse: collapse;border-left-color: #000000;border-left-style: 
    solid;border-left-width: 1px;border-right-color: #000000;border-right-style: solid;border-right-width: 
    1px;border-top-color: #000000;border-top-style: solid;border-top-width: 1px;font-size: Nonepx;font-weight: 
    bold;height: 19pt;text-align: center">PR开启天数</td> <td id="open_pull_requests_statistics!G1" 
    style="border-bottom-color: #000000;border-bottom-style: solid;border-bottom-width: 1px;border-collapse: 
    collapse;border-left-color: #000000;border-left-style: solid;border-left-width: 1px;border-right-color: 
    #000000;border-right-style: solid;border-right-width: 1px;border-top-color: #000000;border-top-style: 
    solid;border-top-width: 1px;font-size: Nonepx;font-weight: bold;height: 19pt;text-align: center">审查者</td> </tr> 
    <tr> <td id="open_pull_requests_statistics!A2" rowspan="20" style="border-bottom-color: 
    #000000;border-bottom-style: solid;border-bottom-width: 1px;border-collapse: collapse;border-left-color: 
    #000000;border-left-style: solid;border-left-width: 1px;border-right-color: #000000;border-right-style: 
    solid;border-right-width: 1px;border-top-color: #000000;border-top-style: solid;border-top-width: 1px;font-size: 
    11.0px;height: 19pt;text-align: center">{0}</td> <td id="open_pull_requests_statistics!B2" 
    style="border-bottom-color: #000000;border-bottom-style: solid;border-bottom-width: 1px;border-collapse: 
    collapse;border-left-color: #000000;border-left-style: solid;border-left-width: 1px;border-right-color: 
    #000000;border-right-style: solid;border-right-width: 1px;border-top-color: #000000;border-top-style: 
    solid;border-top-width: 1px;font-size: 11.0px;height: 19pt">{1}</td> <td id="open_pull_requests_statistics!C2" 
    style="border-bottom-color: #000000;border-bottom-style: solid;border-bottom-width: 1px;border-collapse: 
    collapse;border-left-color: #000000;border-left-style: solid;border-left-width: 1px;border-right-color: 
    #000000;border-right-style: solid;border-right-width: 1px;border-top-color: #000000;border-top-style: 
    solid;border-top-width: 1px;font-size: 11.0px;height: 19pt">{2}</td> <td id="open_pull_requests_statistics!D2" 
    style="border-bottom-color: #000000;border-bottom-style: solid;border-bottom-width: 1px;border-collapse: 
    collapse;border-left-color: #000000;border-left-style: solid;border-left-width: 1px;border-right-color: 
    #000000;border-right-style: solid;border-right-width: 1px;border-top-color: #000000;border-top-style: 
    solid;border-top-width: 1px;font-size: 11.0px;height: 19pt">{3}</td> <td id="open_pull_requests_statistics!E2" 
    style="border-bottom-color: #000000;border-bottom-style: solid;border-bottom-width: 1px;border-collapse: 
    collapse;border-left-color: #000000;border-left-style: solid;border-left-width: 1px;border-right-color: 
    #000000;border-right-style: solid;border-right-width: 1px;border-top-color: #000000;border-top-style: 
    solid;border-top-width: 1px;font-size: 11.0px;height: 19pt;text-align: center">{4}</td> <td 
    id="open_pull_requests_statistics!F2" style="border-bottom-color: #000000;border-bottom-style: 
    solid;border-bottom-width: 1px;border-collapse: collapse;border-left-color: #000000;border-left-style: 
    solid;border-left-width: 1px;border-right-color: #000000;border-right-style: solid;border-right-width: 
    1px;border-top-color: #000000;border-top-style: solid;border-top-width: 1px;font-size: 11.0px;height: 
    19pt;text-align: center">{5}</td> <td id="open_pull_requests_statistics!G2" style="border-bottom-color: 
    #000000;border-bottom-style: solid;border-bottom-width: 1px;border-collapse: collapse;border-left-color: 
    #000000;border-left-style: solid;border-left-width: 1px;border-right-color: #000000;border-right-style: 
    solid;border-right-width: 1px;border-top-color: #000000;border-top-style: solid;border-top-width: 1px;font-size: 
    11.0px;height: 19pt">{6}</td> </tr> </table> </body> </html>
    """.format(sig, repo, ref, link, status, duration, reviewers.replace('"', ''))
    f2 = open(html_file, 'w', encoding='utf-8')
    f2.write(template)
    f2.close()


def excel_optimization(filepath, items_count):
    """
    Adjust styles of the xlsx file
    :param filepath: path of the xlsx file
    :param items_count: count of the raws without the first row
    """
    if not filepath.endswith('.xlsx'):
        return
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    # add borders
    border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))
    for row in ws.rows:
        for cell in row:
            cell.border = border
    # align center
    alignment_center = Alignment(horizontal='center', vertical='center')
    for row in ws.rows:
        row[1].alignment = alignment_center
        row[5].alignment = alignment_center
        row[6].alignment = alignment_center
    # fill for the Duration
    cells = ws.iter_rows(min_row=2, min_col=7, max_col=7)
    yellow_fill = PatternFill("solid", start_color='FFFF00')
    first_stage_fill = PatternFill('solid', start_color='FFDAB9')
    second_stage_fill = PatternFill('solid', start_color='FF7F50')
    third_stage_fill = PatternFill('solid', start_color='FF4500')
    for i in cells:
        value = int(i[0].value)
        if 7 < value <= 30:
            i[0].fill = first_stage_fill
        elif 30 < value <= 365:
            i[0].fill = second_stage_fill
        elif value > 365:
            i[0].fill = third_stage_fill
    # fill for the status mark
    status = ws.iter_rows(min_row=2, min_col=6, max_col=6)
    for j in status:
        value = j[0].value
        if value != '待合入':
            j[0].fill = yellow_fill
    # merge cells
    ws.merge_cells(start_row=2, end_row=items_count + 1, start_column=1, end_column=1)
    # delete auxiliary column
    ws.delete_cols(1)
    wb.save(filepath)
    wb.close()
    # generate html file by the xlsx file
    html_file = filepath.replace('.xlsx', '.html')
    if items_count > 1:
        xlsx2html(filepath, html_file)
    else:
        generate_one_row_table(filepath)
    log.logger.info('Generate {}'.format(html_file))


def send_email(sig, xlsx_file, receivers, host, port, username, password):
    """
    Send email to reviewers
    :param sig: sig name
    :param xlsx_file: path of the xlsx file
    :param receivers: where send to
    :param host: SMTP host
    :param port: SMTP port
    :param username: SMTP username
    :param password: SMTP password
    """
    msg = MIMEMultipart()
    if sig:
        html_file = xlsx_file.replace('.xlsx', '.html')
        with open(html_file, 'r', encoding='utf-8') as f:
            body_of_email = f.read()
        body_of_email = body_of_email.replace('<body>',
                                              '<body><p>下表是SIG {} 所有开启PR的统计，请相关审查者及时跟进</p>'.format(sig)). \
            replace('&nbsp;', '0')
        content = MIMEText(body_of_email, 'html', 'utf-8')
        msg.attach(content)
        msg['Subject'] = '{}的PR汇总'.format(sig)
    else:
        body_of_email = '各SIG的PR处理情况统计见附件'
        content = MIMEText(body_of_email, 'plain', 'utf-8')
        msg.attach(content)
        file = MIMEApplication(open(xlsx_file, 'rb').read())
        file.add_header('Content-Disposition', 'attachment', filename=xlsx_file)
        msg.attach(file)
        msg['Subject'] = '各SIG的PR处理情况统计'

    msg['From'] = username
    msg['To'] = ','.join(receivers)
    try:
        server = smtplib.SMTP(host, int(port))
        server.ehlo()
        server.starttls()
        server.login(username, password)
        server.sendmail(username, receivers, msg.as_string())
        log.logger.info('Sent report email to {}, receivers: {}'.format(sig, receivers))
    except smtplib.SMTPException as e:
        log.logger.error(e)
        sys.exit(1)


def fill_status(status, insert_string):
    """
    Change status of the Pull Request
    :param status: a string of current status
    :param insert_string: abnormal status waiting to add
    :return: status
    """
    if status == '待合入':
        status = insert_string
    else:
        status += '、{}'.format(insert_string)
    return status


def clean_env(data_dir):
    """
    Remove the temporary data
    :param data_dir: directory waiting to clean
    """
    subprocess.call('rm -rf {}'.format(data_dir), shell=True)


def get_repos_pulls_mapping(token, page, enterprise_pulls):
    """
    Get mappings between repos and pulls
    :param token: access token
    :param page: the first page
    :param enterprise_pulls: a list to store pulls
    :return: a dict of {repo: pulls}
    """
    log.logger.info("=" * 25 + " Get enterprise pulls: page {} ".format(page) + "=" * 25)
    url = 'https://gitee.com/api/v5/enterprise/open_euler/pull_requests'
    params = {
        'access_token': token,
        'state': 'open',
        'sort': 'created',
        'direction': 'asc',
        'page': page,
        'per_page': 100
    }
    r = requests.get(url, params=params)
    if r.status_code != 200:
        log.logger.error('Fail to get enterprise pulls list.')
        return
    else:
        enterprise_pulls += r.json()
        if len(r.json()) == 100:
            get_repos_pulls_mapping(token, page + 1, enterprise_pulls)
    return {x['html_url'].split('/', 3)[3]: x for x in enterprise_pulls}


def pr_statistics(data_dir, sigs, repos_pulls_mapping, args):
    """
    :param data_dir: directory to store temporary data
    :param repos_pulls_mapping: mappings between repos and pulls
    :param sigs: a dict of every sig and its repositories
    :param args: arguments passed through the command line
    """
    username = args.username
    port = args.port
    host = args.host
    password = args.password
    log.logger.info('=' * 25 + ' STATISTICS ' + '=' * 25)
    email_mappings = get_email_mappings()
    open_pr_info = []
    for sig in sigs:
        sig_name = sig['name']
        sig_repos = sig['repositories']
        log.logger.info('\nStarting to search sig {}'.format(sig_name))
        if not sig_repos:
            log.logger.info('Find no repositories in sig {}, skip'.format(sig_name))
            continue
        statistics_csv = '{}/statistics_{}.csv'.format(data_dir, sig_name)
        f = codecs.open(statistics_csv, 'w', encoding='utf-8')
        writer = csv.writer(f)
        writer.writerow(['SIG组', '仓库', '目标分支', 'PR链接', 'PR状态', 'PR开启天数', '审查者'])
        maintainers, sig_info_mark = get_maintainers(sig_name)
        receivers = []
        receivers_addrs = []
        open_pr_count = 0
        for full_repo in sig_repos:
            if full_repo.split('/')[0] not in ['src-openeuler', 'openeuler']:
                continue
            open_pr_list = []
            for mapping_key in repos_pulls_mapping.keys():
                if mapping_key.startswith(full_repo + '/'):
                    open_pr_list.append(repos_pulls_mapping[mapping_key])
                    log.logger.info('Find open pr: {}'.format(mapping_key))
            if not open_pr_list:
                continue
            members = maintainers
            if sig_info_mark:
                committers_mapping = get_committers_mapping(sig_name)
                members = get_repo_members(maintainers, committers_mapping, full_repo)
            for member in members:
                if member not in receivers:
                    receivers.append(member)
            for item in open_pr_list:
                link = item['html_url']
                created_at = item['created_at']
                draft = item['draft']
                labels = [x['name'] for x in item['labels']]
                ref_branch = item['base']['ref']
                status = '待合入'
                if draft:
                    status = fill_status(status, '草稿')
                if 'openeuler-cla/yes' not in labels:
                    status = fill_status(status, 'CLA认证失败')
                if 'ci_failed' in labels:
                    status = fill_status(status, '门禁检查失败')
                if not item['mergeable']:
                    status = fill_status(status, '存在冲突')
                if 'kind/wait_for_update' in labels:
                    status = fill_status(status, '等待更新')
                duration = count_duration(created_at)
                open_pr_info.append([sig_name, full_repo, ref_branch, link, status, duration, ','.join(members)])
                link = "<a href='{0}'>{0}</a>".format(link)
                new_members = ["<a href='https://gitee.com/{0}'>{0}</a>".format(x) for x in members]
                writer.writerow([sig_name, full_repo, ref_branch, link, status, duration, ','.join(new_members)])
                open_pr_count += 1
        f.close()
        for receiver in receivers:
            email_address = email_mappings[receiver]
            if email_address and email_address not in receivers_addrs:
                receivers_addrs.append(email_address)
        if receivers_addrs:
            log.logger.info('Ready to send to {}, email addresses: {}'.format(sig_name, receivers_addrs))
            statistics_xlsx = csv_to_xlsx(statistics_csv)
            excel_optimization(statistics_xlsx, open_pr_count)
            send_email(sig_name, statistics_xlsx, receivers_addrs, host, port, username, password)
        else:
            log.logger.warning('Find these receivers all has no email addresses: {}'.format(receivers))
    # generate statistics.csv
    statistics_csv = 'statistics.csv'
    f = codecs.open(statistics_csv, 'w', encoding='utf-8')
    writer = csv.writer(f)
    writer.writerow(['SIG组', '仓库', '目标分支', 'PR链接', 'PR状态', 'PR开启天数', '审查者'])
    for item in open_pr_info:
        writer.writerow(item)
    if args.clean:
        clean_env(data_dir)


def main():
    """
    main function
    """
    parser = argparse.ArgumentParser()
    parser.add_argument('-t', '--token', help='access token', required=True)
    parser.add_argument('-u', '--username', help='SMTP username', required=True)
    parser.add_argument('-p', '--port', help='SMTP port', required=True)
    parser.add_argument('-host', '--host', help='SMTP host', required=True)
    parser.add_argument('-pswd', '--password', help='SMTP password', required=True)
    parser.add_argument('-c', '--clean', help='clean current directory')
    args = parser.parse_args()
    data_dir = prepare_env()
    sigs = get_sigs()
    repos_pulls_mapping = get_repos_pulls_mapping(args.token, 1, [])
    pr_statistics(data_dir, sigs, repos_pulls_mapping, args)


if __name__ == '__main__':
    main()
